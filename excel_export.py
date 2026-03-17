#!/usr/bin/env python3
"""
Excel Export & Databook Generator

Two modes:

  /export-all  →  Quick structured dump of all iterations into one Excel file.
                  No LLM needed. Summary + Findings + per-iteration detail sheets.

  /databook    →  One professional databook PER successful iteration.
                  Each workbook contains:
                    • "Data" sheet — full raw dataset from workspace/data.xlsx
                    • Analysis sheets — with Excel FORMULAS referencing the Data
                      sheet so every number is traceable to the source rows.
                  Generated one at a time via LLM (one call per iteration).

Usage:
  python excel_export.py                # export-all mode
  python excel_export.py --databook     # generate per-iteration databooks
  python excel_export.py --databook -n 3  # only iteration 3

  # Or from orchestrator interactive mode:
  /export-all
  /databook
"""

import json
import logging
import os
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import anthropic
import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv

load_dotenv()
API_KEY = os.getenv("API_KEY")

os.chdir(Path(__file__).parent)

# ─── Logging Setup ────────────────────────────────────────────────────────────

logger = logging.getLogger("ddanalyze.excel_export")


def setup_logging(debug: bool = False, log_dir: str = "logs") -> Path:
    Path(log_dir).mkdir(exist_ok=True)
    log_file = Path(log_dir) / f"excel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    console_level = logging.DEBUG if debug else logging.INFO
    root = logging.getLogger("ddanalyze")
    root.setLevel(logging.DEBUG)
    if not root.handlers:
        ch = logging.StreamHandler(sys.stdout)
        ch.setLevel(console_level)
        ch.setFormatter(logging.Formatter("%(asctime)s %(message)s", datefmt="%H:%M:%S"))
        root.addHandler(ch)
        fh = logging.FileHandler(log_file, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(logging.Formatter(
            "%(asctime)s [%(levelname)-7s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S",
        ))
        root.addHandler(fh)
    return log_file


# ─── File Utilities ───────────────────────────────────────────────────────────

def read_file(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def write_file(path: str, content: str) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)


# ─── Archive Parser ──────────────────────────────────────────────────────────

_INTERNAL_ERROR_LABELS = {
    "JSON_PARSE_ERROR", "CRITIC_JSON_PARSE_ERROR", "TIMEOUT", "FATAL_ERROR",
    "SYNTH_JSON_PARSE_ERROR",
}


def parse_archive_entries(archive_text: str) -> list:
    """Parse full_archive.txt into structured dicts for all non-internal-error entries."""
    separator = "=" * 80
    entries = []
    for block in archive_text.split(separator):
        block = block.strip()
        if not block or "ITERATION:" not in block:
            continue
        status_match = re.search(r"\nSTATUS:\s*(\S+)", block)
        if not status_match:
            continue
        status = status_match.group(1).strip()
        if status.upper() in _INTERNAL_ERROR_LABELS:
            continue
        entry = {"status": status}
        for field, pattern in [
            ("iteration", r"ITERATION:\s*(\d+)"),
            ("date", r"DATE:\s*(.+)"),
            ("analysis_type", r"ANALYSIS TYPE:\s*(.+)"),
            ("hypothesis", r"HYPOTHESIS:\s*(.+)"),
            ("columns_used", r"COLUMNS USED:\s*(.+)"),
        ]:
            m = re.search(pattern, block)
            if m:
                entry[field] = m.group(1).strip()
        dash_sep = "-" * 80
        code_m = re.search(r"CODE:\n(.*?)\n" + re.escape(dash_sep), block, re.DOTALL)
        if code_m:
            entry["code"] = code_m.group(1).strip()
        output_m = re.search(
            r"OUTPUT:\n(.*?)(?:\n" + re.escape(dash_sep) + r"|\Z)", block, re.DOTALL
        )
        if output_m:
            entry["output"] = output_m.group(1).strip()
        eval_m = re.search(r"EVALUATION:\n(.*?)(?:\Z)", block, re.DOTALL)
        if eval_m:
            entry["evaluation"] = eval_m.group(1).strip()
        if "iteration" in entry:
            entries.append(entry)
    return entries


def _parse_evaluation_fields(eval_text: str) -> dict:
    """Extract structured fields from the evaluation text block."""
    fields = {}
    if not eval_text:
        return fields
    quality_m = re.search(r"Quality:\s*(.+)", eval_text)
    if quality_m:
        fields["quality"] = quality_m.group(1).strip()
    summary_m = re.search(
        r"Summary:\s*(.+?)(?=\nKey findings:|\nSuggested followup:|\nError type:|\Z)",
        eval_text, re.DOTALL,
    )
    if summary_m:
        fields["summary"] = summary_m.group(1).strip()
    findings = re.findall(r"  - (.+)", eval_text)
    if findings:
        fields["key_findings"] = findings
    followup_m = re.search(
        r"Suggested followup:\s*(.+?)(?=\nConfirmed dead ends:|\Z)", eval_text, re.DOTALL,
    )
    if followup_m:
        fields["suggested_followup"] = followup_m.group(1).strip()
    error_m = re.search(r"Error type:\s*(.+)", eval_text)
    if error_m:
        fields["error_type"] = error_m.group(1).strip()
    return fields


# ─── LLM Utilities ───────────────────────────────────────────────────────────

def call_llm(
    client: anthropic.Anthropic,
    system: str,
    user: str,
    model: str,
    max_tokens: int,
    tag: str = "LLM",
) -> str:
    logger.debug(f"[{tag}] Sending request | system={len(system):,}ch user={len(user):,}ch")
    t0 = time.time()
    with client.messages.stream(
        model=model, max_tokens=max_tokens, system=system,
        messages=[{"role": "user", "content": user}],
    ) as stream:
        response = stream.get_final_message()
    elapsed = time.time() - t0
    in_tok = response.usage.input_tokens
    out_tok = response.usage.output_tokens
    logger.info(f"  [{tag}] Done in {elapsed:.1f}s | tokens in={in_tok:,} out={out_tok:,}")
    for block in response.content:
        if block.type == "text":
            return block.text
    return ""


def parse_json_response(text: str) -> Optional[dict]:
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    match = re.search(r"```(?:json)?\s*([\s\S]*?)```", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1).strip())
        except json.JSONDecodeError:
            pass
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        try:
            return json.loads(text[start : end + 1])
        except json.JSONDecodeError:
            pass
    return None


# ─── Excel Styling Constants ─────────────────────────────────────────────────

# Deloitte-inspired palette
DARK_GREEN = "046A38"
GREEN = "86BC25"
LIGHT_GREEN = "C4D600"
BLACK = "000000"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
MID_GREY = "D9D9D9"

HEADER_FILL = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color=DARK_GREEN)
SUBTITLE_FONT = Font(name="Arial", bold=True, size=11, color=BLACK)
BODY_FONT = Font(name="Arial", size=10, color="333333")
NUM_FONT = Font(name="Arial", size=10, color="333333")
ACCENT_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
ACCENT_FONT = Font(name="Arial", bold=True, color=WHITE, size=10)
ZEBRA_FILL = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FAILURE_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color=MID_GREY), right=Side(style="thin", color=MID_GREY),
    top=Side(style="thin", color=MID_GREY), bottom=Side(style="thin", color=MID_GREY),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_cell(ws, row: int, col: int, wrap: bool = False) -> None:
    cell = ws.cell(row=row, column=col)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP_ALIGNMENT if wrap else TOP_ALIGNMENT


def _auto_width(ws, max_width: int = 50, min_width: int = 10) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                max_line = max(len(line) for line in lines) if lines else 0
                max_len = max(max_len, max_line)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _style_data_sheet(ws, n_rows: int, n_cols: int) -> None:
    """Apply professional styling to the Data sheet: header + zebra stripes + auto-filter."""
    _style_header_row(ws, 1, n_cols)
    for row_idx in range(2, n_rows + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL
    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    # Freeze header row
    ws.freeze_panes = "A2"
    _auto_width(ws, max_width=30, min_width=8)


# ─── Export-All: Quick Structured Dump ────────────────────────────────────────

def export_iterations_to_excel(
    archive_path: str,
    context_path: str,
    output_path: str,
    graphs_folder: str = "workspace/graphs",
) -> str:
    """
    Export all iterations to a single Excel workbook (no LLM needed).
    Sheets: Summary, Findings, per-iteration details, Graphs Index.
    """
    archive_text = read_file(archive_path)
    entries = parse_archive_entries(archive_text)
    if not entries:
        logger.warning("[Export] No entries found in archive.")
        return ""

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_path = output_path if output_path.endswith(".xlsx") else \
        os.path.join(output_path, f"iterations_export_{timestamp}.xlsx")
    Path(final_path).parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"[Export] Writing {len(entries)} iterations to {final_path}")

    with pd.ExcelWriter(final_path, engine="openpyxl") as writer:
        # Sheet 1: Summary
        summary_rows = []
        for e in entries:
            ef = _parse_evaluation_fields(e.get("evaluation", ""))
            summary_rows.append({
                "Iteration": int(e.get("iteration", 0)),
                "Date": e.get("date", ""),
                "Status": e.get("status", ""),
                "Analysis Type": e.get("analysis_type", ""),
                "Hypothesis": e.get("hypothesis", ""),
                "Quality": ef.get("quality", ""),
                "Summary": ef.get("summary", ""),
                "Columns Used": e.get("columns_used", ""),
            })
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
        ws = writer.sheets["Summary"]
        ws.cell(row=1, column=1, value="DDAnalyze — Iteration Summary").font = TITLE_FONT
        _style_header_row(ws, 2, len(df_summary.columns))
        for r in range(3, len(df_summary) + 3):
            for c in range(1, len(df_summary.columns) + 1):
                _style_data_cell(ws, r, c, wrap=(c >= 5))
            sv = ws.cell(row=r, column=3).value
            if sv and str(sv).lower() == "success":
                ws.cell(row=r, column=3).fill = SUCCESS_FILL
            elif sv:
                ws.cell(row=r, column=3).fill = FAILURE_FILL
        _auto_width(ws)

        # Sheet 2: Findings
        findings_rows = []
        for e in entries:
            if e.get("status", "").lower() != "success":
                continue
            ef = _parse_evaluation_fields(e.get("evaluation", ""))
            for f in ef.get("key_findings", []):
                findings_rows.append({
                    "Iteration": int(e.get("iteration", 0)),
                    "Analysis Type": e.get("analysis_type", ""),
                    "Finding": f,
                })
            if ef.get("suggested_followup"):
                findings_rows.append({
                    "Iteration": int(e.get("iteration", 0)),
                    "Analysis Type": e.get("analysis_type", ""),
                    "Finding": f"[FOLLOWUP] {ef['suggested_followup']}",
                })
        if findings_rows:
            df_f = pd.DataFrame(findings_rows)
            df_f.to_excel(writer, sheet_name="Findings", index=False, startrow=1)
            ws2 = writer.sheets["Findings"]
            ws2.cell(row=1, column=1, value="Key Findings & Follow-ups").font = TITLE_FONT
            _style_header_row(ws2, 2, len(df_f.columns))
            for r in range(3, len(df_f) + 3):
                for c in range(1, len(df_f.columns) + 1):
                    _style_data_cell(ws2, r, c, wrap=(c == 3))
            _auto_width(ws2)

        # Per-iteration detail sheets
        for e in entries:
            it = int(e.get("iteration", 0))
            sn = f"Iter_{it:02d}"[:31]
            ef = _parse_evaluation_fields(e.get("evaluation", ""))
            rows = [
                {"Field": "Iteration", "Value": it},
                {"Field": "Date", "Value": e.get("date", "")},
                {"Field": "Status", "Value": e.get("status", "")},
                {"Field": "Analysis Type", "Value": e.get("analysis_type", "")},
                {"Field": "Hypothesis", "Value": e.get("hypothesis", "")},
                {"Field": "Columns Used", "Value": e.get("columns_used", "")},
                {"Field": "Quality", "Value": ef.get("quality", "")},
                {"Field": "Summary", "Value": ef.get("summary", "")},
            ]
            for i, finding in enumerate(ef.get("key_findings", []), 1):
                rows.append({"Field": f"Finding {i}", "Value": finding})
            if ef.get("suggested_followup"):
                rows.append({"Field": "Suggested Followup", "Value": ef["suggested_followup"]})
            if ef.get("error_type"):
                rows.append({"Field": "Error Type", "Value": ef["error_type"]})
            rows.append({"Field": "Code", "Value": e.get("code", "")})
            rows.append({"Field": "Output", "Value": e.get("output", "")})
            df_d = pd.DataFrame(rows)
            df_d.to_excel(writer, sheet_name=sn, index=False, startrow=1)
            wsd = writer.sheets[sn]
            wsd.cell(row=1, column=1,
                     value=f"Iteration {it} — {e.get('analysis_type', '')}").font = TITLE_FONT
            _style_header_row(wsd, 2, 2)
            for r in range(3, len(df_d) + 3):
                _style_data_cell(wsd, r, 1)
                _style_data_cell(wsd, r, 2, wrap=True)
                wsd.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10, color="333333")
            wsd.column_dimensions["A"].width = 20
            wsd.column_dimensions["B"].width = 100

        # Graphs Index
        graphs_dir = Path(graphs_folder)
        if graphs_dir.exists():
            pngs = sorted(graphs_dir.glob("*.png"))
            if pngs:
                gr = []
                for p in pngs:
                    im = re.match(r"iter(\d+)", p.stem)
                    gr.append({
                        "Iteration": int(im.group(1)) if im else 0,
                        "Filename": p.name,
                        "Size (KB)": round(p.stat().st_size / 1024, 1),
                        "Path": str(p),
                    })
                df_g = pd.DataFrame(gr)
                df_g.to_excel(writer, sheet_name="Graphs", index=False, startrow=1)
                wsg = writer.sheets["Graphs"]
                wsg.cell(row=1, column=1, value="Generated Graphs Index").font = TITLE_FONT
                _style_header_row(wsg, 2, len(df_g.columns))
                for r in range(3, len(df_g) + 3):
                    for c in range(1, len(df_g.columns) + 1):
                        _style_data_cell(wsg, r, c)
                _auto_width(wsg)

    size_kb = Path(final_path).stat().st_size / 1024
    logger.info(f"[Export] SUCCESS — {final_path} ({size_kb:.1f} KB, {len(entries)} iterations)")
    return final_path


# ─── Per-Iteration Databook Generator ─────────────────────────────────────────
#
# For each successful iteration we produce one standalone .xlsx:
#   Sheet "Data"             — full raw dataset, styled, with auto-filter & freeze
#   Sheet(s) for analysis    — LLM-generated openpyxl code that builds analysis
#                              tables using Excel FORMULAS (SUMIFS, COUNTIFS, etc.)
#                              referencing the "Data" sheet, so every number is
#                              traceable back to the source rows.
#

DATABOOK_SYSTEM_PROMPT = """You are an expert Excel data engineer producing a professional financial
due diligence databook. You will receive:

1. The DATA SHEET SCHEMA — column names, their Excel column letters (A, B, C…),
   data types, and sample values. The raw data occupies sheet "Data" rows 2..N+1
   (row 1 = headers). Total data rows = N.
2. ONE analysis iteration — its type, hypothesis, the Python code that was run,
   and the printed output with the results/tables.

Your job: write Python code using openpyxl that OPENS an existing workbook
(already containing the "Data" sheet) and ADDS one or more analysis sheets.

═══════════════════════════════════════════════════════════════════════
CRITICAL RULES — READ CAREFULLY
═══════════════════════════════════════════════════════════════════════

A) FORMULAS, NOT VALUES
   Every number in your analysis sheets MUST come from an Excel formula that
   references the "Data" sheet. Use:
     - SUMIFS, COUNTIFS, AVERAGEIFS for aggregations
     - SUMPRODUCT for weighted calcs, conditional counts, unique counts
     - INDEX/MATCH or VLOOKUP for lookups
     - Simple arithmetic (+, -, /, *) to combine formula cells
   NEVER hard-code numeric results. The user must be able to change a number
   in the Data sheet and see the analysis update automatically.

B) REFERENCE FORMAT
   Use the pattern:  Data!$C$2:$C${n+1}  where n = total data rows (provided).
   Always anchor column references with $ (e.g., Data!$C$2:$C$101).
   The variable `n` (total data rows) will be passed to your code.

C) STRUCTURE
   - Your code receives: wb (the open workbook), n (total data rows), col_map (dict of
     column_name → Excel letter, e.g. {"Revenue": "D", "Client": "B"}).
   - Create sheets with ws = wb.create_sheet("Sheet Name")
   - Write headers, then formulas.
   - Do NOT touch the "Data" sheet (it's already complete).
   - Do NOT save or close the workbook — the caller handles that.

D) PROFESSIONAL STYLING
   Apply clean, professional formatting:
   - Use openpyxl.styles (Font, PatternFill, Border, Alignment, numbers)
   - Dark green headers (#046A38, white bold text)
   - Zebra striping on data rows (light grey #F2F2F2 on even rows)
   - Thin grey borders (#D9D9D9)
   - Number format "#,##0" for integers, "#,##0.00" for decimals,
     '€#,##0' or '€#,##0.0,"k"' for currency
   - Percentage format "0.0%"
   - Column widths set appropriately
   - Freeze panes on header row
   - Titles in row 1 (merged across columns, bold, 14pt dark green)

E) SHEET NAMING
   Name sheets descriptively (max 31 chars). Examples:
   "Revenue by Year", "Top 20 Clients", "Channel Mix", "Seasonality"

F) RETURN FORMAT
   Return ONLY a JSON object:
   {
     "sheets": ["Sheet Name 1", "Sheet Name 2"],
     "code": "python code as a string"
   }
   The code must define a function: def build_analysis(wb, n, col_map):
   No preamble. No markdown fences around the JSON.

G) KEEP IT FOCUSED
   Only create sheets relevant to THIS specific analysis iteration.
   Typically 1-3 sheets per iteration. Quality over quantity.
"""


def _get_data_schema(data_file: str) -> tuple:
    """
    Read the data file and return:
      - df: the full DataFrame
      - col_map: dict of {column_name: Excel letter}
      - schema_text: human-readable schema description for the LLM
    """
    df = pd.read_excel(data_file)
    col_map = {}
    schema_lines = []
    for i, col in enumerate(df.columns):
        letter = get_column_letter(i + 1)
        col_map[col] = letter
        dtype = str(df[col].dtype)
        sample = str(df[col].iloc[0]) if len(df) > 0 else ""
        sample2 = str(df[col].iloc[min(1, len(df) - 1)]) if len(df) > 1 else ""
        schema_lines.append(f"  Column {letter}: \"{col}\" ({dtype}) — e.g. {sample}, {sample2}")

    schema_text = (
        f"Total rows: {len(df)} (data in rows 2..{len(df)+1}, headers in row 1)\n"
        f"Total columns: {len(df.columns)}\n"
        + "\n".join(schema_lines)
    )
    return df, col_map, schema_text


def _write_data_sheet(wb, df: pd.DataFrame) -> None:
    """Write the raw DataFrame to the 'Data' sheet with professional styling."""
    ws = wb.active
    ws.title = "Data"

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Handle NaN/NaT
            if pd.isna(value):
                cell.value = None
            else:
                cell.value = value

    # Style it
    _style_data_sheet(ws, len(df), len(df.columns))


def _build_iteration_context(entry: dict) -> str:
    """Build the LLM user message for a single iteration."""
    eval_fields = _parse_evaluation_fields(entry.get("evaluation", ""))
    findings = eval_fields.get("key_findings", [])
    findings_str = "\n".join(f"  - {f}" for f in findings) if findings else "(none)"

    return f"""## Analysis Iteration {entry.get('iteration', '?')}

Type: {entry.get('analysis_type', 'unknown')}
Hypothesis: {entry.get('hypothesis', '')}
Columns Used: {entry.get('columns_used', '')}
Status: {entry.get('status', '')}
Quality: {eval_fields.get('quality', '')}

### Key Findings
{findings_str}

### Summary
{eval_fields.get('summary', '')}

### Original Python Code
```python
{entry.get('code', '')}
```

### Printed Output
{entry.get('output', '')}
"""


def generate_single_databook(
    client: anthropic.Anthropic,
    model: str,
    entry: dict,
    data_file: str,
    output_dir: str,
    max_tokens: int = 16000,
) -> str:
    """
    Generate one professional databook for a single successful iteration.

    Creates an .xlsx with:
      - "Data" sheet: full raw dataset with styling, auto-filter, freeze panes
      - Analysis sheets: formula-linked to Data sheet, generated by LLM

    Returns the path of the created file, or empty string on failure.
    """
    iter_num = int(entry.get("iteration", 0))
    analysis_type = entry.get("analysis_type", "unknown")
    safe_type = re.sub(r"[^\w\s-]", "", analysis_type).strip().replace(" ", "_")[:30]

    output_filename = f"databook_iter{iter_num:02d}_{safe_type}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    logger.info(f"\n{'─' * 50}")
    logger.info(f"[Databook] Iteration {iter_num}: {analysis_type}")
    logger.info(f"[Databook] Output: {output_path}")

    # ── Step 1: Read data and build schema ────────────────────────────────
    try:
        df, col_map, schema_text = _get_data_schema(data_file)
    except Exception as e:
        logger.error(f"[Databook] Failed to read data file: {e}")
        return ""

    n = len(df)

    # ── Step 2: Create workbook with Data sheet ───────────────────────────
    from openpyxl import Workbook
    wb = Workbook()
    _write_data_sheet(wb, df)

    # Save the base workbook (Data sheet only) — LLM code will add to it
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    logger.info(f"  [Databook] Data sheet written ({n} rows, {len(df.columns)} cols)")

    # ── Step 3: Call LLM to generate analysis sheets ──────────────────────
    iteration_context = _build_iteration_context(entry)
    col_map_str = json.dumps(col_map, indent=2)

    user_msg = f"""## Data Sheet Schema
{schema_text}

## Column Map (column_name → Excel letter)
{col_map_str}

## Variable n = {n}  (total data rows; data in Data!A2:A{n+1})

{iteration_context}

Write a `build_analysis(wb, n, col_map)` function that adds analysis sheets
to the workbook. Use Excel formulas referencing the Data sheet for ALL numbers.
Keep it focused on this specific analysis — typically 1-3 sheets."""

    response = call_llm(
        client, DATABOOK_SYSTEM_PROMPT, user_msg, model, max_tokens,
        tag=f"Databook-Iter{iter_num}",
    )

    parsed = parse_json_response(response)
    if parsed is None:
        logger.error(f"  [Databook] Failed to parse LLM response for iteration {iter_num}")
        # Still return the file with just the Data sheet
        return output_path

    sheets = parsed.get("sheets", [])
    code = parsed.get("code", "")

    if not code:
        logger.warning(f"  [Databook] LLM returned no code for iteration {iter_num}")
        return output_path

    logger.info(f"  [Databook] Planned sheets: {sheets}")

    # ── Step 4: Execute the LLM-generated code ───────────────────────────
    # Wrap the code in a script that opens the workbook, calls build_analysis, saves
    wrapper_script = f'''import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers

# LLM-generated code
{code}

# Execute
wb = load_workbook("{output_path}")
col_map = json.loads("""{json.dumps(col_map)}""")
n = {n}

try:
    build_analysis(wb, n, col_map)
    wb.save("{output_path}")
    print("SUCCESS: Saved to {output_path}")
    print(f"Sheets: {{[s.title for s in wb.worksheets]}}")
except Exception as e:
    print(f"ERROR: {{e}}", file=sys.stderr)
    # Still try to save what we have
    try:
        wb.save("{output_path}")
        print("Partial save completed", file=sys.stderr)
    except:
        pass
    sys.exit(1)
finally:
    wb.close()
'''

    script_path = "workspace/databook_script.py"
    write_file(script_path, wrapper_script)

    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"

    for attempt in range(1, 3):
        try:
            result = subprocess.run(
                [sys.executable, script_path],
                capture_output=True, text=True, encoding="utf-8",
                errors="replace", env=env, timeout=180,
            )
        except subprocess.TimeoutExpired:
            logger.error(f"  [Databook] Timed out for iteration {iter_num}")
            return output_path  # Return with Data sheet only

        if result.stdout:
            logger.info(f"  [Databook] {result.stdout.strip()}")

        if not result.stderr:
            break  # Success

        if attempt >= 2:
            logger.error(f"  [Databook] Failed after retry:\n{result.stderr[:500]}")
            return output_path  # Return with Data sheet only

        # Ask LLM to fix
        logger.warning(f"  [Databook] Error (attempt {attempt}):\n{result.stderr[:400]}")
        logger.info("  [Databook] Asking LLM to fix...")

        fix_msg = f"""The build_analysis code produced an error. Fix it.

## Error
{result.stderr}

## stdout
{result.stdout[:500] if result.stdout else '(none)'}

## Original Code
```python
{code}
```

## Data Sheet Schema
{schema_text}

## Column Map
{col_map_str}

## n = {n}

Return ONLY a JSON object: {{"sheets": [...], "code": "fixed code"}}"""

        fix_response = call_llm(
            client, DATABOOK_SYSTEM_PROMPT, fix_msg, model, max_tokens,
            tag=f"DatabookFix-Iter{iter_num}",
        )
        fix_parsed = parse_json_response(fix_response)
        if fix_parsed and fix_parsed.get("code"):
            code = fix_parsed["code"]
            wrapper_script = f'''import json
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers

# LLM-generated code (fixed)
{code}

# Execute
wb = load_workbook("{output_path}")
col_map = json.loads("""{json.dumps(col_map)}""")
n = {n}

try:
    build_analysis(wb, n, col_map)
    wb.save("{output_path}")
    print("SUCCESS: Saved to {output_path}")
    print(f"Sheets: {{[s.title for s in wb.worksheets]}}")
except Exception as e:
    print(f"ERROR: {{e}}", file=sys.stderr)
    try:
        wb.save("{output_path}")
    except:
        pass
    sys.exit(1)
finally:
    wb.close()
'''
            write_file(script_path, wrapper_script)
        else:
            logger.error("  [Databook] Failed to parse fix response")
            return output_path

    # Verify and report
    if Path(output_path).exists():
        size_kb = Path(output_path).stat().st_size / 1024
        try:
            check_wb = load_workbook(output_path, read_only=True)
            sheet_names = [s.title for s in check_wb.worksheets]
            check_wb.close()
            logger.info(
                f"  [Databook] DONE — {output_path} ({size_kb:.1f} KB) "
                f"Sheets: {sheet_names}"
            )
        except Exception:
            logger.info(f"  [Databook] DONE — {output_path} ({size_kb:.1f} KB)")
        return output_path

    return ""


def generate_databooks(
    client: anthropic.Anthropic,
    model: str,
    archive_path: str,
    data_file: str,
    output_dir: str,
    max_tokens: int = 16000,
    only_iteration: Optional[int] = None,
) -> list:
    """
    Generate one databook per successful iteration.

    Args:
        only_iteration: If set, only generate for this specific iteration number.

    Returns list of created file paths.
    """
    if not Path(archive_path).exists():
        logger.error("[Databook] No archive found. Run data analysis first.")
        return []

    archive_text = read_file(archive_path)
    entries = parse_archive_entries(archive_text)
    success_entries = [e for e in entries if e.get("status", "").lower() == "success"]

    if not success_entries:
        logger.error("[Databook] No successful analyses in archive.")
        return []

    if only_iteration is not None:
        success_entries = [e for e in success_entries if int(e.get("iteration", 0)) == only_iteration]
        if not success_entries:
            logger.error(f"[Databook] No successful iteration #{only_iteration} found.")
            return []

    logger.info(f"\n{'═' * 60}")
    logger.info(f"DDAnalyze DATABOOK GENERATOR — {len(success_entries)} iteration(s)")
    logger.info(f"Output dir: {output_dir}")
    logger.info(f"{'═' * 60}")

    created = []
    for i, entry in enumerate(success_entries, 1):
        iter_num = entry.get("iteration", "?")
        logger.info(f"\n[Databook] Processing {i}/{len(success_entries)} (Iteration {iter_num})...")

        result = generate_single_databook(
            client, model, entry, data_file, output_dir, max_tokens
        )
        if result:
            created.append(result)

    logger.info(f"\n{'═' * 60}")
    logger.info(f"[Databook] COMPLETE — {len(created)}/{len(success_entries)} databooks created")
    for p in created:
        size_kb = Path(p).stat().st_size / 1024
        logger.info(f"  {Path(p).name} ({size_kb:.1f} KB)")
    logger.info(f"{'═' * 60}")

    return created


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(
        description="DDAnalyze Excel Export & Databook Generator"
    )
    parser.add_argument(
        "--databook", action="store_true",
        help="Generate per-iteration databooks with Data + formula-linked analysis sheets",
    )
    parser.add_argument(
        "--output", "-o", type=str, default=None,
        help="Output path (file for export-all, directory for databook)",
    )
    parser.add_argument(
        "-n", "--iteration", type=int, default=None,
        help="Only generate databook for this iteration number",
    )
    args = parser.parse_args()

    config = yaml.safe_load(read_file("config.yaml"))
    debug_logging = config.get("debug_logging", False)
    log_file = setup_logging(debug=debug_logging)

    archive_path = config.get("archive_file", "full_archive.txt")
    context_path = config.get("active_context_file", "active_context.md")
    data_file = config.get("data_file", "workspace/data.xlsx")
    graphs_folder = config.get("graphs_folder", "workspace/graphs")

    if not Path(archive_path).exists():
        logger.error(f"Archive not found: {archive_path}")
        logger.error("Run data analysis first (python loop.py or python orchestrator.py)")
        sys.exit(1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.databook:
        if not Path(data_file).exists():
            logger.error(f"Data file not found: {data_file}")
            sys.exit(1)

        model = config.get("model", "claude-sonnet-4-6")
        max_tokens = config.get("databook_max_tokens", 16000)
        output_dir = args.output or f"workspace/exports/databooks_{timestamp}"

        client = anthropic.Anthropic(api_key=API_KEY)

        results = generate_databooks(
            client, model, archive_path, data_file, output_dir, max_tokens,
            only_iteration=args.iteration,
        )

        if results:
            print(f"\n{len(results)} databook(s) created in {output_dir}/")
            for p in results:
                print(f"  {Path(p).name}")
        else:
            print("\nNo databooks created. Check the log for details.")
            sys.exit(1)
    else:
        output_path = args.output or f"workspace/exports/iterations_export_{timestamp}.xlsx"

        logger.info(f"\n{'═' * 60}")
        logger.info("DDAnalyze EXCEL EXPORT")
        logger.info(f"Output : {output_path}")
        logger.info(f"Log    : {log_file}")
        logger.info(f"{'═' * 60}")

        result = export_iterations_to_excel(
            archive_path, context_path, output_path, graphs_folder
        )
        if result:
            print(f"\nExport created: {result}")
        else:
            print("\nExport failed. Check the log for details.")
            sys.exit(1)


if __name__ == "__main__":
    main()
